from openpyxl import load_workbook
import csv
import re
import os


def fin_year_str(year):
    return '%d-%d' % (year, (year+1)-2000)


filename = '../data/provincial/from-jonathan/2015 MTEF/NW - EPRE - 2015-16 - Final.xlsm'
wb = load_workbook(filename=filename, read_only=True, data_only=True)
match = re.search('(?P<province_code>[A-Z]{2}) - EPRE - (?P<financial_year>\d{4})-\d{2} - Final.xlsm', filename)
province_code = match.group('province_code')
budget_financial_year = int(match.group('financial_year'))

ws = wb['Settings']

department_sheets = {}

for row in ws.iter_rows(min_row=13, max_row=13+20):
    if row[4].value:
        sheet_name = row[3].value
        department_name = row[4].value.strip()
        department_sheets[sheet_name] = department_name
        print(sheet_name, department_name)

with open('epre.csv', 'w') as csv_file:
    fieldnames = [
        'department',
        'programme',
        'financial_year',
        'phase',
        'economic_classification_1',
        'economic_classification_2',
        'economic_classification_3',
        'amount',
    ]
    writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
    writer.writeheader()
    for sheet_name, department_name in department_sheets.items():
        ws = wb[str(sheet_name)]
        rows = {}

        print("\n", sheet_name, department_name)

        # skip to programmes
        for idx, row in enumerate(ws.iter_rows()):
            if row[1].value == 'PROGRAMME DETAILS':
                programmes_row_idx = idx+2
                break

        econ_classs = {}
        prev_econ_class_level = 0
        in_econ_class = False
        for idx, row in enumerate(ws.iter_rows(min_row=programmes_row_idx)):
            if (row[1].value
                and row[1].font
                and row[1].font.color
                and row[1].font.color.rgb == 'FF0000FF'):
                if row[1].value not in ('Direct Charges'):
                    programme_name = row[1].value.strip()

            if row[1].value == 'Total':
                in_econ_class = True
                continue
            if row[1].value == 'Total economic classification':
                in_econ_class = False

            if in_econ_class:
                econ_class = row[1].value.strip()
                econ_class_level = int(row[1].alignment.indent)

                if econ_class_level < prev_econ_class_level:
                    for key, value in list(econ_classs.items()):
                        if key > econ_class_level:
                            del econ_classs[key]

                econ_classs[econ_class_level] = econ_class

                rows_key = tuple([department_name, programme_name]
                                 + [econ_classs[k]
                                    for k in sorted(econ_classs)])

                # Drop subtotal that includes this row
                if econ_class_level > prev_econ_class_level:
                    print("Removing ", prev_rows_key)
                    del rows[prev_rows_key]

                print("Adding ", rows_key)
                rows[rows_key] = []

                columns = {
                    3: (-4, 'Outcome'),
                    4: (-3, 'Outcome'),
                    5: (-2, 'Outcome'),
                    6: (-1, 'Main appropriation'),
                    7: (-1, 'Adjusted appropriation'),
                    8: (-1, 'Revised estimate'),
                    16: (0, 'Budget'),
                    25: (1, 'MTEF'),
                    27: (2, 'MTEF'),
                }

                for col_idx, col_meta in columns.items():
                    year_offset, phase = col_meta
                    amount = row[col_idx-1].value
                    if amount:
                        amount = round(amount)
                    financial_year = fin_year_str(budget_financial_year+year_offset)
                    rows[rows_key].append({
                        'phase': phase,
                        'financial_year': financial_year,
                        'amount': amount,
                    })

                # End
                prev_rows_key = rows_key
                prev_econ_class_level = econ_class_level

        for row_key, row in rows.items():
            for year_phase in row:
                year_phase['department'] = row_key[0]
                year_phase['programme'] = row_key[1]
                year_phase['economic_classification_1'] = row_key[2]
                if len(row_key) > 3:
                    year_phase['economic_classification_2'] = row_key[3]
                else:
                    year_phase['economic_classification_2'] = None
                if len(row_key) > 4:
                    year_phase['economic_classification_3'] = row_key[4]
                else:
                    year_phase['economic_classification_3'] = None
                writer.writerow(year_phase)
